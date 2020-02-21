import openpyxl, pyperclip, os
from openpyxl.utils import get_column_letter

# ExcelWorksheetを読み込む
wbname = input('Excel-Workbook-Name: ')
wbsheet = input('Excel-Workbook-Sheet-Name: ')

wb = openpyxl.load_workbook(wbname, data_only=True)
sheet = wb[wbsheet]
sheet = wb.active

caption = input('caption: ')
while True:
	linecode = input('linecode: ')
	if len(linecode) == sheet.max_column:
		break
	print('wrong code. input again.')

decimal_places = int(input('decimal places: '))

OutputText_begintable =  '\\begin{table}[h]\n'
OutputText_begintabular = 	'\t\\centering\n' \
							'\t\\begin{tabular}{' + linecode + '} \n' \
                    		'\t\t\\hline\n'
    

OutputText_end =    '\t\t\\hline\n' \
					'\t\\end{tabular}\n' \
                    '\\end{table}\n'

caption_text = '\t\\caption{' + caption + '}\n'

# クリップ用ファイルへの書き込み
clip_file = open('{}.txt'.format(caption), 'w')

clip_file.write(OutputText_begintable)
clip_file.write(caption_text)
clip_file.write(OutputText_begintabular)

max_cell = get_column_letter(sheet.max_column) + str(sheet.max_row)

for row_of_cell_obj in sheet['B2':max_cell]:
    for cell_obj in row_of_cell_obj:
        cell_obj.value = round(cell_obj.value, decimal_places)

for row_of_cell_obj in sheet['A1':max_cell]:
	clip_file.write('\t\t')
	for cell_obj in row_of_cell_obj:
		if cell_obj.column == sheet.max_column:
			clip_file.write(str(cell_obj.value) + '\t\\\\')
		else:
			clip_file.write(str(cell_obj.value) + '\t&\t')
	if row_of_cell_obj == sheet[1]:
		clip_file.write('\n\t\t\\hline \\hline\n')
	else:
		clip_file.write('\n')

clip_file.write(OutputText_end)

print()

# 結果をターミナル上に表示
print('-----------BEGIN OF TEXT-----------')

print(OutputText_begintable, end='')
print(caption_text, end='')
print(OutputText_begintabular, end='')

for row_of_cell_obj in sheet['A1':max_cell]:
	print('\t\t', end='')
	for cell_obj in row_of_cell_obj:
		if cell_obj.column == sheet.max_column:
			print(cell_obj.value, end='\t\\\\')
		else:
			print(cell_obj.value, end='\t&\t')
	if row_of_cell_obj == sheet[1]:
		print('\n\t\t\\hline \\hline')
	else:
		print()

print(OutputText_end, end='')

print('------------END OF TEXT------------')

clip_file.close()

# クリップ用テキストの中身を読み込む
clip_file = open('{}.txt'.format(caption), 'r')
copy_clip = clip_file.read()
clip_file.close()

# クリップボードにコピー
pyperclip.copy(copy_clip)

# クリップ用ファイルを削除
os.remove(caption + '.txt')